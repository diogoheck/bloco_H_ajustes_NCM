from audioop import lin2ulaw
from pdb import line_prefix
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

PASTA_DE_TRABALHO = 'Bloco_H.xlsx'
PLANILHA_0200 = '0200'
PLANILHA_H010 = 'H010'
PLANILHA_H020 = 'H020'
PLANLHA_SPED = 'SPED'
PLANILHA_NF_CREDITO = 'NF_CREDITO'

CABECALHO_0200 = ['', 'REG', 'COD_ITEM', 'DESCR_ITEM', 'COD_BARRA',
                  'COD_ANT_ITEM', 'UNID_INV', 'TIPO_ITEM',
                  'COD_NCM', 'EX_IPI', 'COD_GEN', 'COD_LST',
                  'ALIQ_ICMS', 'CEST', 'NCM_VALIDA', 'C100', '1923', 'H010', 'K200']
CABECALHO_H010 = ['', 'REG', 'COD_ITEM', 'UNID', 'QTD',
                  'VLR_UNIT', 'VL_ITEM', 'IND_PROP',
                  'COD_PART', 'TXT_COMPL', 'COD_CTA',
                  'VL_ITEM_IR']

CABECALHO_H020 = ['', 'REG', 'CST_ICMS', 'BC_ICMS',
                  'VL_ICMS']

CABECALHO_NF_CREDITO = ['REG', 'COD_ITEM',
                        'COD_NCM', 'DESCR_ITEM', 'QTD',
                        'VLR_UNIT', 'BASE COM MARGEM',
                        'ALIQ_ICMS', 'VL_ICMS',
                        'CREDITO_TOTAL']


class QTDE:
    linhas_0200 = 0
    linhas_H010 = 0
    linhas_H020 = 0
    linhas_SPED_0200 = 0


def aplicar_soma(ws, tamanho_plan):
    for i in range(2, tamanho_plan + 2):

        ws['C' + str(i)].number_format = '#,##0.00'
        ws['D' + str(i)].number_format = '#,##0.00'
        ws['E' + str(i)] = f'=sum(C{i}:D{i})'
        ws['E' + str(i)].number_format = '#,##0.00'


def aplicar_procv_h020(ws, tamanho_plan):

    for i in range(3, tamanho_plan + 3):

        # aliquota icms
        ws['F' +
            str(i)] = f'=VLOOKUP(A{i},H010!$A$2:$M${QTDE.linhas_H010 + 1},6,FALSE)'

        # base ICMS
        ws['E' +
            str(i)] = f'=sum(product(D{i},J1),D{i})'

        # valor ICMS
        ws['G' +
            str(i)] = f'=round(product(product(e{i},f{i}),0.01), 2)'

        # ws['F' + str(i)] = f'=quociente(f{i},100)'

        ws['D' + str(i)].number_format = '#,##0.00'
        ws['E' + str(i)].number_format = '#,##0.00'
        ws['G' + str(i)].number_format = '#,##0.00'
        # ws['F' + str(i)].number_format = '0%'


# def concatenar_0200(ws_0200, i):




#     A = str(ws_0200['A' + str(i)].value) if ws_0200['A' + str(i)].value else ''
#     B = str(ws_0200['B' + str(i)].value) if ws_0200['B' + str(i)].value else ''
#     C = str(ws_0200['C' + str(i)].value) if ws_0200['C' + str(i)].value else ''
#     D = str(ws_0200['D' + str(i)].value) if ws_0200['D' + str(i)].value else ''
#     E = str(ws_0200['E' + str(i)].value) if ws_0200['E' + str(i)].value else ''
#     F = str(ws_0200['F' + str(i)].value) if ws_0200['F' + str(i)].value else ''
#     G = str(ws_0200['G' + str(i)].value) if ws_0200['G' + str(i)].value else ''
#     H = str(ws_0200['H' + str(i)].value) if ws_0200['H' + str(i)].value else ''
#     I = str(ws_0200['I' + str(i)].value) if ws_0200['I' + str(i)].value else ''
#     J = str(ws_0200['J' + str(i)].value) if ws_0200['J' + str(i)].value else ''
#     K = str(ws_0200['K' + str(i)].value) if ws_0200['K' + str(i)].value else ''
#     L = str(ws_0200['L' + str(i)].value).replace('.',
#                                                  ',') if ws_0200['L' + str(i)].value else ''
#     M = str(ws_0200['M' + str(i)].value) if ws_0200['M' + str(i)].value else ''

#     concatenar = '|' + A + '|' + B + '|' + C +  \
#                  '|' + D + '|' + E + '|' + F +  \
#                  '|' + G + '|' + H + '|' + I +  \
#                  '|' + J + '|' + K + '|' + L +  \
#                  '|' + M + '|'

#     return concatenar


def concatenar_0220(ws_0200, i):
    A = str(ws_0200['A' + str(i)].value) if ws_0200['A' + str(i)].value else ''
    B = str(ws_0200['B' + str(i)].value) if ws_0200['B' + str(i)].value else ''
    C = str(ws_0200['C' + str(i)].value) if ws_0200['C' + str(i)].value else ''

    return '|' + A + '|' + B + '|' + C + '||'


def converte_valor_para_texto(texto):
    if '.' in texto and ',' in texto:
        valor = texto.replace(',', '').replace('.', ',')
    else:
        valor = texto.replace('.', ',')
    return valor


def concatenar_h010(ws_H010, i):

    concatenar = f'=concatenate("|", H010!b{i}, "|", H010!c{i}, "|" ,H010!g{i}, \
                                "|", H010!h{i}, "|", H010!i{i}, "|", H010!j{i}, "|", \
                                H010!k{i}, "|", H010!l{i}, "|", H010!m{i}, "|", \
                                H010!n{i}, "|", H010!o{i}, "|")'

    return concatenar

def concatenar_0200(ws_sped, i):

    concatenar = f'=concatenate("|", 0200!a{i}, "|", 0200!b{i}, "|" ,0200!c{i}, \
                                "|", 0200!d{i}, "|", 0200!e{i}, "|", 0200!f{i}, "|", \
                                0200!g{i}, "|", 0200!h{i}, "|", 0200!i{i}, "|", \
                                0200!j{i}, "|", 0200!k{i}, "|", 0200!l{i}, "|", 0200!m{i}, "|")'

    return concatenar


def concatenar_h020(ws_H020, i):

    concatenar = f'=concatenate("|", H020!b{i}, "|", H020!c{i}, "|" ,H020!e{i}, "|" ,H020!g{i}, "|")'
    return concatenar


def buscar_registros_h010_h020(ws_sped, ws_H010, ws_H020, tamanho_plan_h010):
    linha_sped = QTDE.linhas_SPED_0200 + 1
    for i in range(2, tamanho_plan_h010 + 2):
        ws_sped['A' + str(linha_sped)] = concatenar_h010(ws_H010, i)
        linha_sped += 1
        ws_sped['A' + str(linha_sped)] = concatenar_h020(ws_H020, i + 1)
        linha_sped += 1


def buscar_registros_0200(ws_sped, ws_0200, tamanho_plan_0200):
    j = 0
    # Percorrer 0200
    for i in range(2, tamanho_plan_0200 + 2):
        if (ws_0200['N' + str(i)].value == 'S' and ws_0200['Q' + str(i)].value == 'S') or \
            ws_0200['O' + str(i)].value == 'S' or \
                ws_0200['P' + str(i)].value == 'S' or ws_0200['R' + str(i)].value == 'S':
            j = j + 1
            QTDE.linhas_SPED_0200 += 1

            ws_sped['A' + str(j)] = concatenar_0200(ws_0200, i)
        if ws_0200['A' + str(i)].value == '0220':
            j = j + 1
            QTDE.linhas_SPED_0200 += 1
            ws_sped['A' + str(j)] = concatenar_0220(ws_0200, i)

    QTDE.linhas_SPED_0200 += 1


def aplicar_procv_nf_credito(ws, tamanho_plan):

    for i in range(1, tamanho_plan + 1):

        # REG
        ws['A' +
            str(i + 1)] = i

        # CODIGO ITEM
        ws['B' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H010!$A$2:$M${QTDE.linhas_H010 + 1},3,FALSE)'
        # NCM
        ws['C' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H010!$A$2:$M${QTDE.linhas_H010 + 1},5,FALSE)'
        # DESCRIÇÃO
        ws['D' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H010!$A$2:$M${QTDE.linhas_H010 + 1},4,FALSE)'
        # QUANTIDADE
        ws['E' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H010!$A$2:$M${QTDE.linhas_H010 + 1},8,FALSE)'
        # VALOR UNITARIO (TOTAL)
        ws['F' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H010!$A$2:$M${QTDE.linhas_H010 + 1},9,FALSE)'
        # BASE UNITARIA H020
        ws['G' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H020!$A$3:$M${QTDE.linhas_H020 + 2},5,FALSE)'
        # ALIQUOTA
        ws['H' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H020!$A$3:$M${QTDE.linhas_H020 + 2},6,FALSE)'
        # CREDITO UNITARIO 020
        ws['I' +
            str(i + 1)] = f'=VLOOKUP(A{i + 1},H020!$A$3:$M${QTDE.linhas_H020 + 2},7,FALSE)'
        # CREDIDO TOTAL

        # ws['F' + str(i)] = f'=quociente(f{i},100)'

        ws['C' + str(i + 1)].number_format = '#,##0.00'
        ws['D' + str(i + 1)].number_format = '#,##0.00'
        ws['E' + str(i + 1)].number_format = '#,##0.00'
        ws['F' + str(i + 1)].number_format = '#,##0.00'
        ws['G' + str(i + 1)].number_format = '#,##0.00'
        ws['H' + str(i + 1)].number_format = '#,##0.00'
        ws['I' + str(i + 1)].number_format = '#,##0.00'
        ws['J' + str(i + 1)].number_format = '#,##0.00'
        # ws['F' + str(i)].number_format = '0%'
        ws['J' +
            str(i + 1)] = f'=product(e{i + 1},i{i + 1})'
    ws['L' + str(1)] = 'CREDITO_TOTAL'
    ws['L' + str(2)] = f'=sum(j{2}:J{i+1})'
    ws['L' + str(1)].font = Font(size='14', bold=True)
    ws['L' + str(2)].font = Font(size='14', bold=True)
    ws['L' + str(1)].number_format = '#,##0.00'
    ws['L' + str(2)].number_format = '#,##0.00'


def aplicar_procv_h010(ws, tamanho_plan):

    for i in range(2, tamanho_plan + 2):

        # DESCRIÇÃO
        ws['D' +
            str(i)] = f'=VLOOKUP(C{i},0200!$B$2:$M${QTDE.linhas_0200 + 1},2,FALSE)'
        # NCM
        ws['E' +
            str(i)] = f'=VLOOKUP(C{i},0200!$B$2:$M${QTDE.linhas_0200 + 1},7,FALSE)'
        # ALIQUOTA
        ws['F' +
            str(i)] = f'=VLOOKUP(C{i},0200!$B$2:$M${QTDE.linhas_0200 + 1},11,FALSE)'

        ws['I' + str(i)].number_format = '#,##0.00'
        ws['J' + str(i)].number_format = '#,##0.00'
        ws['O' + str(i)].number_format = '#,##0.00'


def formata_numero(ws, intervalo):
    columns = ws[intervalo]

    for rows in columns:
        for cell in rows:
            cell.number_format = '#,##0.00'


def autosize(ws):
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True


def salvar_0200(PASTA_DE_TRABALHO, dic_0200, PLANILHA_0200):
    wb = load_workbook(PASTA_DE_TRABALHO)
    ws = wb[PLANILHA_0200]
    ws.append(CABECALHO_0200)
    for k, v in dic_0200.items():
        # for lista in lista_0200:
        QTDE.linhas_0200 += 1
        ws.append(v)

    ws.delete_cols(1)
    autosize(ws)
    wb.save(PASTA_DE_TRABALHO)


def salvar_h010(PASTA_DE_TRABALHO, lista_h010_h020, PLANILHA_H010):
    wb = load_workbook(PASTA_DE_TRABALHO)
    ws = wb[PLANILHA_H010]
    ws.append(CABECALHO_H010)

    for lista in lista_h010_h020:
        for li in lista:
            if li[1] == 'H010':
                QTDE.linhas_H010 += 1
                li[0] = QTDE.linhas_H010
                ws.append(li)

    # ws.delete_cols(1)
    ws.insert_cols(4, 3)
    # cabeçalho adicional
    ws['D1'].value = 'DESCR_ITEM'
    ws['E1'].value = 'COD_NCM'
    ws['F1'].value = 'ALIQ_ICMS'
    aplicar_procv_h010(ws, QTDE.linhas_H010)
    autosize(ws)
    wb.save(PASTA_DE_TRABALHO)


def salvar_h020(PASTA_DE_TRABALHO, lista_h010_h020, PLANILHA_H020):
    wb = load_workbook(PASTA_DE_TRABALHO)
    ws = wb[PLANILHA_H020]
    ws.append(CABECALHO_H020)

    for lista in lista_h010_h020:
        for li in lista:
            if li[1] == 'H020':
                QTDE.linhas_H020 += 1
                li[0] = QTDE.linhas_H020
                ws.append(li)

    # ws.delete_cols(1)
    ws.insert_cols(5, 2)
    ws.insert_rows(1)
    ws['I1'].value = 'MARGEM'
    ws['J1'].value = 0
    ws['J1'].number_format = '0%'
    ws['E2'].value = 'BASE COM MARGEM'
    ws['F2'].value = 'ALIQ_ICMS'
    aplicar_procv_h020(ws, QTDE.linhas_H020)

    autosize(ws)
    wb.save(PASTA_DE_TRABALHO)


def salvar_NF_CREDITO(PASTA_DE_TRABALHO, PLANILHA_NF_CREDITO):
    wb = load_workbook(PASTA_DE_TRABALHO)
    ws = wb[PLANILHA_NF_CREDITO]
    ws.append(CABECALHO_NF_CREDITO)

    aplicar_procv_nf_credito(ws, QTDE.linhas_H020)

    autosize(ws)
    wb.save(PASTA_DE_TRABALHO)


def salvar_sped_0200(PASTA_DE_TRABALHO, PLANLHA_SPED, PLANILHA_0200):
    wb = load_workbook(PASTA_DE_TRABALHO)
    ws_sped = wb[PLANLHA_SPED]
    ws_0200 = wb[PLANILHA_0200]
    # ws.append(CABECALHO_NF_CREDITO)

    buscar_registros_0200(ws_sped, ws_0200, QTDE.linhas_0200)

    autosize(ws_sped)
    autosize(ws_0200)
    wb.save(PASTA_DE_TRABALHO)


def salvar_sped_h010_h020(PASTA_DE_TRABALHO, PLANLHA_SPED,
                          PLANILHA_H010, PLANILHA_H020):
    wb = load_workbook(PASTA_DE_TRABALHO)
    ws_sped = wb[PLANLHA_SPED]
    ws_H010 = wb[PLANILHA_H010]
    ws_H020 = wb[PLANILHA_H020]

    buscar_registros_h010_h020(ws_sped, ws_H010, ws_H020, QTDE.linhas_H010)

    autosize(ws_sped)
    wb.save(PASTA_DE_TRABALHO)


if __name__ == '__main__':
    salvar_0200(PASTA_DE_TRABALHO, 'XX', PLANILHA_H010)
