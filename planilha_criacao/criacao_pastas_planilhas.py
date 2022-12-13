from openpyxl import Workbook
from openpyxl import load_workbook

PASTA_DE_TRABALHO = 'Bloco_H.xlsx'
PLANILHA_0200 = '0200'
PLANILHA_H010 = 'H010'
PLANILHA_H020 = 'H020'
PLANLHA_SPED = 'SPED'
PLANILHA_NF_CREDITO = 'NF_CREDITO'


def criacao_pasta_trabalho(PASTA_DE_TRABALHO):
    wb = Workbook()
    wb.save(PASTA_DE_TRABALHO)


def criar_nome_planilha(ws, nome):
    if ws.title != 'Sheet':
        return True
    else:
        ws.title = nome
        return False


def criacao_planilhas(*plans):
    wb = load_workbook(PASTA_DE_TRABALHO)
    ws = wb.active
    for nome_plan in plans:
        if criar_nome_planilha(ws, nome_plan):
            ws = wb.create_sheet(nome_plan)
    wb.save(PASTA_DE_TRABALHO)


if __name__ == '__main__':
    criacao_pasta_trabalho(PASTA_DE_TRABALHO)
    criacao_planilhas(PLANILHA_0200, PLANILHA_H010,
                      PLANILHA_H020, PLANILHA_NF_CREDITO, PLANLHA_SPED)
    print('Pasta de trabalho criada com sucesso!')
